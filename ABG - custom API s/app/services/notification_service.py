import json
from flask import request
import requests
from datetime import datetime
from app.config import AWS_SECRET_KEY, AWS_REGION, AWS_ACCESS_KEY, AWS_BUCKET_NAME
from app import mongo
from bson.objectid import ObjectId
from openpyxl import load_workbook
import os
from app.utils.api_call import make_api_call
from app.utils.s3 import download_file_from_s3


def send_notification(data, token):
    if data:
        notification_type = data.get('notification_type', None)
        if notification_type and notification_type in ['event', 'general']:
            description = data.get('description', None)
            if description:
                if notification_type == 'general':
                    employee = data.get('employee', None)
                    if employee:
                        object_key = data.get('employee', [{}])[0].get('path')
                        parts = object_key.split('/')
                        object_key = '/'.join(parts[3:])
                        file_data = download_file_from_s3(AWS_BUCKET_NAME, object_key, AWS_ACCESS_KEY, AWS_SECRET_KEY,
                                                          AWS_REGION)
                        status_code = employee_bulk_upload(notification_type, description, object_key, token)
                        if status_code != 200:
                            return 400, 'Something wrong with file upload', {}
                    else:
                        return 400, 'Invalid employee file', {}
                else:
                    event_id = data.get('event_id', None)
                    if event_id:
                        data_code, msg = prepare_event_notify_data(event_id, description, token)
                        if data_code != 200:
                            return 400, 'Something wrong with notification data', {}
                    else:
                        return 400, 'Event id is missing', {}
                return 200, "Successfully inserted", {}
            else:
                return 400, 'Invalid description', {}
        else:
            return 400, 'notification type should be either event or general', {}
    else:
        return 400, 'Invalid input', {}


def employee_bulk_upload(notification_type, description, object_key, token):
    cpath = os.getcwd()
    filename = cpath + '/' + object_key
    # Open the workbook
    workbook = load_workbook(filename)
    sheet = workbook.active

    if sheet:
        # checking for columns
        required_columns = ['NAME', 'CONTACT NUMBER', 'EMAIL ID']
        header_row = next(sheet.iter_rows(min_row=1, max_row=1))
        header_values = [cell.value for cell in header_row]
        if header_values:
            missing_columns = [column for column in required_columns if column not in header_values]
            if missing_columns:
                return 400, f'Invalid Headers : {", ".join(missing_columns)}', {}
        else:
            return 400, 'Something went wrong', {}

        # Find rows with errors
        error_rows = []
        correct_rows = []
        is_error = False
        for idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):  # Start from row 2 to skip header
            if all(cell.value is None for cell in row):
                # Skip empty rows
                # If it's the last row, don't skip it
                if idx == sheet.max_row:
                    break
                continue

            # Validate the row for errors
            result, error_data = employee_data_has_error(row)
            if not result:
                error_rows.append(error_data)
            else:
                correct_rows.append(row)
        if len(correct_rows) != 0:
            # Retrieve data from the file when there are no errors
            data_for_notify = prepare_employee_data(correct_rows, description, token)

            if data_for_notify:
                for user_data in data_for_notify:
                    # insert user details into notification tbl
                    notification_id = mongo.db.notification.insert_one(user_data).inserted_id

        os.remove(filename)
        return 200
    else:
        return 400


def employee_data_has_error(row):
    name = row[0].value
    contact_number = row[1].value
    email = row[2].value

    error_flag = False
    errors = []

    if not name or not isinstance(name, str):
        print("Name is missing or not a string.")
        error_flag = True
        errors.append((row, "Name is missing or not a string."))

    if not contact_number or not isinstance(contact_number, int):
        print("Contact is missing")
        error_flag = True
        errors.append((row, "Contact number is missing."))

    if not email or not isinstance(email, str):
        print("Contact is missing")
        error_flag = True
        errors.append((row, "Contact number is missing."))

    if error_flag:
        return False, errors
    else:
        return True, None


def prepare_employee_data(correct_rows, description, token):
    data_for_notify = []
    for row in correct_rows:
        url = os.environ.get('USER_DETAILS_URL') + str(row[2].value)
        headers = {'x-access-token': token}
        response_post = make_api_call(url=url, headers=headers)
        if response_post['status_code'] == 200:
            detail = {}
            detail['user_id'] = ObjectId(response_post['data']['_id'])
            detail['name'] = str(response_post['data']['contact_details']['first_name'])
            detail['email'] = str(response_post['data']['contact_details']['email'])
            detail['description'] = description
            detail['notification_type'] = 'general'
            detail['datetime'] = str(datetime.now())
            data_for_notify.append(detail)
    return data_for_notify


def prepare_event_notify_data(event_id, description, token):
    user_ids = [ObjectId(notification['user_id']) for notification in
                mongo.db.notification.find({'event_id': ObjectId(event_id)})]
    print(user_ids)
    if user_ids:
        for user in user_ids:
            payload = {}
            payload_json = json.dumps(payload)
            url = os.environ.get('USER_AUTH_URL')
            headers = {'x-access-token': token, 'Content-Type': 'application/json'}
            response_post = requests.request("POST", url, headers=headers, data=payload_json)
            if response_post.status_code == 200:
                current_user = response_post.json()
                detail = {}
                detail['user_id'] = ObjectId(user)
                detail['event_id'] = ObjectId(event_id)
                detail['name'] = str(current_user['data']['contact_details']['first_name'])
                detail['email'] = str(current_user['data']['contact_details']['email'])
                detail['description'] = description
                detail['notification_type'] = 'event'
                detail['datetime'] = str(datetime.now())
                notification_id = mongo.db.notification.insert_one(detail).inserted_id
        return 200, ' success'
    else:
        return 400, 'something went wrong !'


def notification_list(current_user):
    current_user_id =current_user['data']['_id']
    if current_user['data']['role'] in ['initiator',"employee"] :
        notification_list = list(mongo.db.notification.find({'user_id':ObjectId(current_user_id)}))
    else:
        notification_list = list(mongo.db.notification.find({}))
    notifications = []
    if notification_list:
        for data in notification_list:
            data['_id'] = str(data['_id'])
            data['user_id'] = str(data['user_id'])
            if 'event_id' in data and data['event_id']:
                data['event_id'] = str(data['event_id'])
            notifications.append(data)
        return 200, "Notification List", notifications
    else:
        return 400, "No data found", {}

