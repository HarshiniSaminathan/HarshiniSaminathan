from flask import jsonify, request
from openpyxl import load_workbook, Workbook
from app import mongo, config
from jsonschema import validate, exceptions, ValidationError
import json
from datetime import datetime
import os
from bson.objectid import ObjectId
from app.utils.api_call import make_api_call
from app.utils.helpers import schema_validator
from app.utils.s3 import download_file_from_s3


def employee_bulk_upload(event_id, object_key, event_name,token):
    cpath = os.getcwd()
    filename = cpath + '/' + object_key
    # Open the workbook
    workbook = load_workbook(filename)
    sheet = workbook.active

    if sheet:
        # checking for columns
        required_columns = ['NAME', 'CONTACT NUMBER', 'EMAIL ID']
        header_row = next(sheet.iter_rows(min_row=1, max_row=1))
        header_values = [cell.value for cell in header_row]
        if header_values:
            missing_columns = [column for column in required_columns if column not in header_values]
            if missing_columns:
                return 400, f'Invalid Headers : {", ".join(missing_columns)}', {}
        else:
            return 400, 'Something went wrong', {}

        # Find rows with errors
        error_rows = []
        correct_rows = []
        is_error = False
        for idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):  # Start from row 2 to skip header
            if all(cell.value is None for cell in row):
                # Skip empty rows
                # If it's the last row, don't skip it
                if idx == sheet.max_row:
                    break
                continue

            # Validate the row for errors
            result, error_data = employee_data_has_error(row)
            if not result:
                error_rows.append(error_data)
            else:
                correct_rows.append(row)
        if len(correct_rows) != 0:
            # Retrieve data from the file wh, 'success'en there are no errors
            data_for_notify, employee_data = prepare_employee_data(correct_rows, event_id, event_name,token)
            if data_for_notify:
                for user_data in data_for_notify:
                    # insert user details into notification tbl
                    notification_id = mongo.db.notification.insert_one(user_data).inserted_id

            if employee_data:
                mongo.db.workflow_instance.update_one({'_id': ObjectId(event_id)},
                                                      {'$set': {'employee': employee_data}})

        # Delete the downloaded file
        os.remove(filename)
        return 200
    else:
        return 400


def employee_data_has_error(row):
    name = row[0].value
    contact_number = row[1].value
    email = row[2].value

    error_flag = False
    errors = []

    if not name or not isinstance(name, str):
        print("Name is missing or not a string.")
        error_flag = True
        errors.append((row, "Name is missing or not a string."))

    if not contact_number or not isinstance(contact_number, int):
        print("Contact is missing")
        error_flag = True
        errors.append((row, "Contact number is missing."))

    if not email or not isinstance(email, str):
        print("Contact is missing")
        error_flag = True
        errors.append((row, "Contact number is missing."))

    if error_flag:
        return False, errors
    else:
        return True, None


def prepare_employee_data(correct_rows, event_id, event_name,token):
    data_for_notify = []
    employee_data = []
    for row in correct_rows:  # Start from row 2 to skip header
        # checking that employee is existing or not through make_api_call
        url = os.environ.get('USER_DETAILS_URL') + str(row[2].value)
        headers = {'x-access-token':token}
        response_post = make_api_call(url=url,headers=headers)
        if response_post['status_code'] == 200:
            detail = {}
            detail['event_id'] = event_id
            detail['user_id'] = ObjectId(response_post['data']['_id'])
            detail['name'] = str(response_post['data']['contact_details']['first_name'])
            detail['email'] = str(response_post['data']['contact_details']['email'])
            detail['description'] = 'Hi there, Inviting you to attend ' + event_name
            detail['notification_type'] = 'event'
            detail['datetime'] = str(datetime.now())
            data_for_notify.append(detail)

            employee = {}
            employee['user_id'] = ObjectId(response_post['data']['_id'])
            employee['name'] = str(response_post['data']['contact_details']['first_name'])
            employee['email'] = str(response_post['data']['contact_details']['email'])
            employee_data.append(employee)
    return data_for_notify, employee_data
