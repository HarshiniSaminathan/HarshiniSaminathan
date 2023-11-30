import json
import os

import openpyxl
from flask import request, send_file
import base64
import pandas as pd
import re

from werkzeug.utils import secure_filename

from app.controller.patientController import findPatientId
from app.response import failure_response, success_response
from app.controller.userController import check_email_existence
from app.controller.adminController import (insert_doctor,insert_role_password,updateSlots,fetch_doctor_records,addFeedbackResponse,get_feedbacks,fetch_patient_records,
                                            get_total_doctor,insert_admin,fetch_admin_records,get_total_admin,insert_slot,check_slot_inserted)


def register_doctor():
    try:
        data = request.get_json()
        required_fields = ['doctorName', 'doctorPhoneNumber', 'doctorAddress', 'doctorExperience',
                            'doctorSpecialization', 'doctorEmailId', 'doctorSpecializationProof', 'password']
        for field in required_fields:
            if field not in data or not data[field]:
                return failure_response(statuscode='400', content=f'Missing or empty field: {field}')

        doctorName = data['doctorName']
        doctorPhoneNumber = data['doctorPhoneNumber']
        doctorAddress = data['doctorAddress']
        doctorExperience = data['doctorExperience']
        doctorSpecialization = data['doctorSpecialization']
        doctorEmailId = data['doctorEmailId']
        doctorSpecializationProof = data['doctorSpecializationProof']
        password = data['password']
        role = "DOCTOR"
        if check_email_existence(doctorEmailId):
            return failure_response(statuscode='409', content='Email id already exists')
        try:
            binary_data = base64.b64decode(doctorSpecializationProof)
        except Exception as e:
            return failure_response(statuscode='400', content='Invalid base64 encoding for doctorSpecializationProof')
        print(f"Encoded data: {doctorSpecializationProof}")
        print(f"Decoded data: {binary_data}")
        insert_role_password(doctorEmailId, password, role)
        insert_doctor(doctorName, doctorPhoneNumber, doctorAddress, doctorExperience,
                      doctorSpecialization, binary_data, doctorEmailId)
        return success_response('Doctor Added Successfully')

    except Exception as e:
        print(f"Error: {e}")
        return failure_response(statuscode='500', content='An unexpected error occurred.')


def get_Register_Doctor_Records():
    try:
        page_header = request.headers.get('Page')
        per_page_header = request.headers.get('Per-Page')
        if not page_header:
            return failure_response(statuscode='401', content='page_header is missing')
        if not per_page_header:
            return failure_response(statuscode='401', content='per_page_header is missing')

        doctor_info, total_pages = fetch_doctor_records(int(page_header), int(per_page_header))
        total_doctors = get_total_doctor()
        return success_response({'data': doctor_info, 'Doctor-Total-Count': str(total_doctors), 'Pagination': str(total_pages)})
    except Exception as e:
        print(f"Error: {e}")
        return failure_response(statuscode='500', content='An unexpected error occurred.')


def register_Admin():
    try:
        data = request.get_json()
        required_fields = ['adminName', 'adminPhoneNumber', 'adminAddress', 'emailId',
                            'password']
        for field in required_fields:
            if field not in data or not data[field]:
                return failure_response(statuscode='400', content=f'Missing or empty field: {field}')

        adminName = data['adminName']
        adminPhoneNumber = data['adminPhoneNumber']
        adminAddress = data['adminAddress']
        emailId  = data['emailId']
        password= data['password']
        role = "ADMIN"
        if check_email_existence(emailId):
            return failure_response(statuscode='409', content='Email id already exists')
        insert_role_password(emailId, password, role)
        insert_admin(adminName, adminPhoneNumber, adminAddress, emailId)
        return success_response('Admin Added Successfully')
    except Exception as e:
        print(f"Error: {e}")
        return failure_response(statuscode='500', content='An unexpected error occurred.')


def get_Register_Admin_Records():
    try:
        page_header = request.headers.get('Page')
        per_page_header = request.headers.get('Per-Page')
        if not page_header:
            return failure_response(statuscode='401', content='page_header is missing')
        if not per_page_header:
            return failure_response(statuscode='401', content='per_page_header is missing')
        data,total_pages = fetch_admin_records(int(page_header), int(per_page_header))
        total_admin = get_total_admin()
        return success_response({'data': data, 'Admin-Total-Count': str(total_admin),'Pagination':total_pages})
    except Exception as e:
        print(f"Error: {e}")
        return failure_response(statuscode='500', content='An unexpected error occurred.')


def add_Slot_To_Doctors():
    try:
        data = request.get_json()
        required_fields = ['doctorEmailId','slotStartTime','slotEndTime']
        for field in required_fields:
            if field not in data or not data[field]:
                return failure_response(statuscode='400', content=f'Missing or empty field: {field}')
        doctorEmailId = data['doctorEmailId']
        slotStartTime = data['slotStartTime']
        slotEndTime = data['slotEndTime']
        slotStatus = True    #  initially slotstatus will be True
        if check_email_existence(doctorEmailId):
            if check_slot_inserted(doctorEmailId,slotStartTime):
                return failure_response(statuscode='409', content='Ths slot timimgs inserted already')

            insert_slot(doctorEmailId, slotStartTime, slotStatus, slotEndTime)
            return success_response('Slot Added Successfully')

        return failure_response(statuscode='409', content='Email id does not exists')
    except Exception as e:
        print(f"Error: {e}")
        return failure_response(statuscode='500', content='An unexpected error occurred.')


def update_Slots_status(doctorEmailId):
    try:
        if check_email_existence(doctorEmailId):
            try:
                data = request.get_json()
                slotStatus = data['slotStatus']
                slotStartTime = data['slotStartTime']
                slotEndTime = data['slotEndTime']
                if check_email_existence(doctorEmailId):
                    updateSlots(doctorEmailId,slotStartTime, slotEndTime,slotStatus)
                    return success_response('Slots updated successfully')
                return failure_response(statuscode='409', content='Email id does not exists')
            except Exception as e:
                print(f"Error: {e}")
                return failure_response(statuscode='500', content=str(e))
        else:
            return failure_response(statuscode='500',content='Emailid Does Not Exists')
    except Exception as e:
        print(f"Error: {e}")
        return failure_response(statuscode='500', content='An unexpected error occurred.')


def response_For_Feedback():
    try:
        data=request.get_json()
        required_fields = ['patientEmailId', 'feedbackText', 'rating','feedbackResponse']
        for field in required_fields:
            if field not in data or not data[field]:
                return failure_response(statuscode='400', content=f'Missing or empty field: {field}')
        patientEmailId = data['patientEmailId']
        feedbackText = data['feedbackText']
        rating = data['rating']
        feedbackResponse=data['feedbackResponse']
        try:
            patientId=findPatientId(patientEmailId)
            addFeedbackResponse(patientId,feedbackText,rating,feedbackResponse)
            return success_response('FeedbackResponse added successfully')
        except Exception as e:
            print(f"Error: {e}")
            return failure_response(statuscode='500', content=str(e))
    except Exception as e:
        print(f"Error: {e}")
        return failure_response(statuscode='500', content='An unexpected error occurred.')


def get_All_Feedback():
    try:
        page_header = request.headers.get('Page')
        per_page_header = request.headers.get('Per-Page')
        if not page_header:
            return failure_response(statuscode='401', content='page_header is missing')
        if not per_page_header:
            return failure_response(statuscode='401', content='per_page_header is missing')
        try:
            datas,total_page=get_feedbacks(int(page_header), int(per_page_header))
            if datas:
                return success_response({"data":datas,"Pagination":total_page})
            else:
                return success_response({"data":None})
        except Exception as e:
            print(f"Error: {e}")
            return failure_response(statuscode='500', content=str(e))
    except Exception as e:
        print(f"Error: {e}")
        return failure_response(statuscode='500', content='An unexpected error occurred.')

file_name = None
def uploading_Doctor_Excel():
    global file_name
    try:
        from run import UPLOAD_FOLDER
        file = request.files['file']
        if file:
            filename = secure_filename(file.filename)
            file_name=filename
            file.save(os.path.join(UPLOAD_FOLDER, filename))
            print(file_name)
            return success_response("Excel Uploaded Successfully")
        return failure_response(statuscode='409', content='File Not Found')
    except Exception as e:
        print(f"Error: {e}")
        return failure_response(statuscode='500', content='An unexpected error occurred.')

import xlsxwriter
def download_Errors_InExcel():
    from run import UPLOAD_FOLDER
    global file_name
    try:
        filepath = os.path.join(UPLOAD_FOLDER, file_name)
        if filepath:
            original_data = read_excel_data(filepath)
            errors,validData = validate_data(original_data)

            if errors:
                errors_filepath = os.path.join(UPLOAD_FOLDER, 'errors_output.xlsx')
                write_errors_to_excel(errors_filepath, errors)

                return send_file(errors_filepath, as_attachment=True)
            else:
                return {'data': 'All data is valid'}
        else:
            return {'statuscode': '500', 'content': 'File Not Found'}

    except Exception as e:
        return {'statuscode': '500', 'content': str(e)}


def read_excel_data(filepath):
    import openpyxl

    wb = openpyxl.load_workbook(filepath)
    sheet = wb.active
    original_data = []
    header = ['S_NO', 'DOCTOR_NAME', 'DOCTOR_EMAILID', 'DOCTOR_PHONENUMBER', 'DOCTOR_ADDRESS', 'DOCTOR_SPECIALIZATION']

    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_data = dict(zip(header, row))
        original_data.append(row_data)

    return original_data


def write_errors_to_excel(errors_filepath, errors):
    workbook = xlsxwriter.Workbook(errors_filepath)
    worksheet = workbook.add_worksheet()

    header = ['S_NO', 'DOCTOR_NAME', 'DOCTOR_EMAILID', 'DOCTOR_PHONENUMBER', 'DOCTOR_ADDRESS', 'DOCTOR_SPECIALIZATION',
              'ERRORS']
    for col_num, col_name in enumerate(header):
        worksheet.write(0, col_num, col_name)

    for row_num, error_row in enumerate(errors, start=1):
        for col_num, col_value in enumerate(error_row.values()):
            worksheet.write(row_num, col_num, col_value)
    workbook.close()


def validate_data(data):
    errors = []
    validData=[]
    for row in data:
        error_dict = {}
        error = ""

        if not str(int(row['S_NO'])).isdigit():
            error = error + 'S.No should contain only numbers.'
            error_dict['S_NO'] = row['S_NO']
        error_dict['S_NO'] = row['S_NO']

        if not str(row['DOCTOR_NAME']).replace(" ", "").isalpha():
            error = error + 'Doctor name should not contain numbers or special characters.'
            error_dict['DOCTOR_NAME'] = row['DOCTOR_NAME']
        error_dict['DOCTOR_NAME'] = row['DOCTOR_NAME']

        if not re.match(r"[^@]+@[^@]+\.[^@]+", str(row['DOCTOR_EMAILID'])):
            error = error + 'Invalid email format.'
            error_dict['DOCTOR_EMAILID'] = row['DOCTOR_EMAILID']
        error_dict['DOCTOR_EMAILID'] = row['DOCTOR_EMAILID']

        if not re.match(r'^\d+(\.\d+)?$', str(row['DOCTOR_PHONENUMBER'])):
            error = error + 'Invalid phone number format.'
            error_dict['DOCTOR_PHONENUMBER'] = row['DOCTOR_PHONENUMBER']
        error_dict['DOCTOR_PHONENUMBER'] = row['DOCTOR_PHONENUMBER']

        if not re.match(r'^\d+,[\s\S]+-\d{6}$', str(row['DOCTOR_ADDRESS'])):
            error = error + 'Invalid address format.'
            error_dict['DOCTOR_ADDRESS'] = row['DOCTOR_ADDRESS']
        error_dict['DOCTOR_ADDRESS'] = row['DOCTOR_ADDRESS']

        if not str(row['DOCTOR_SPECIALIZATION']).isalpha():
            error = error + 'Specialization should contain alphabets only.'
            error_dict['DOCTOR_SPECIALIZATION'] = row['DOCTOR_SPECIALIZATION']
        error_dict['DOCTOR_SPECIALIZATION'] = row['DOCTOR_SPECIALIZATION']

        if error != "":
            error_dict['ERRORS'] = error

        if error_dict and error != "":
            errors.append({**row, **error_dict})

        if error == "":
            validData.append({**row, **error_dict})

    return errors,validData


def download_Valid_InExcel():
    from run import UPLOAD_FOLDER
    global file_name
    try:
        filepath = os.path.join(UPLOAD_FOLDER, file_name)
        if filepath:
            original_data = read_excel_data(filepath)

            errors, validData = validate_data(original_data)

            if validData:
                errors_filepath = os.path.join(UPLOAD_FOLDER, 'Valid_output.xlsx')
                write_errors_to_excel(errors_filepath, validData)

                return send_file(errors_filepath, as_attachment=True)
            else:
                return success_response({'data': 'All data is valid'})
        else:
            return failure_response(statuscode='500', content='File Not Found')

    except Exception as e:
        return {'statuscode': '500', 'content': str(e)}


def patient_Excel():
    from run import UPLOAD_FOLDER
    try:
        doctor_info = fetch_patient_records()
        if doctor_info:
            patient_filepath = os.path.join(UPLOAD_FOLDER, 'Patient_Info.xlsx')
            patientInfo_to_excel( doctor_info,patient_filepath)

            return send_file(patient_filepath, as_attachment=True)
            # return success_response({'data': doctor_info})
        else:
            return failure_response(statuscode='409', content='File Not Found')
    except Exception as e:
        print(f"Error: {e}")
        return failure_response(statuscode='500', content='An unexpected error occurred.')


def patientInfo_to_excel(data, excel_filepath):
    workbook = xlsxwriter.Workbook(excel_filepath)
    worksheet = workbook.add_worksheet()
    headers = ["patientFirstName", "patientLastName", "patientPhoneNumber", "patientDOB", "patientAddress", "patientEmailId"]
    for col_num, header in enumerate(headers):
        worksheet.write(0, col_num, header)

    for row_num, patient in enumerate(data, start=1):
        for col_num, key in enumerate(headers):
            worksheet.write(row_num, col_num, patient.get(key, ""))

    workbook.close()