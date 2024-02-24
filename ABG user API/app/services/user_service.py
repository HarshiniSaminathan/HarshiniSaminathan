from datetime import datetime
from datetime import timedelta
from app.utils.jwt_auth import get_jwt
from app.utils.general_utils import generate_session_code, create_new_session, send_email
from flask import g, request
from app import mongo, bcrypt, config
from app.utils.helpers import schema_validator
from bson.objectid import ObjectId
from openpyxl import load_workbook, Workbook
import jwt
import os
# import datetime
from app.utils.api_call import make_api_call
from app.utils.general_utils import is_valid_email, is_valid_phone_number


def user_login(data):
    device_id = data.get('device_id')
    email = data['login_details']['email']
    password = data['login_details']['password']
    # existing_device = mongo.db.device.find_one({'_id': ObjectId(device_id)})
    # if existing_device:
    if email and password:
        existing_user = mongo.db.user.find_one(
            {'contact_details.email': email, 'password': password})
        if existing_user:
            session_code = create_new_session(user_id=existing_user['_id'], device_id=device_id)
            if session_code != None:
                g.session_code = session_code
                mongo.db.user.update_one(
                    {'contact_details.email': email, 'password': password},
                    {'$set': {'session_code': session_code,'logout_at':"",'is_active':True}}
                )
                token = get_jwt(existing_user['_id'], existing_user['org_id'], session_code)
                user_details = {'user_id': str(existing_user['_id']),
                                'role': existing_user['role'], 'org_type': existing_user['org_type'],
                                'org_id': existing_user['org_id'],
                                'status': existing_user['status'],
                                'first_name': existing_user['contact_details']['first_name'],
                                'last_name': existing_user['contact_details']['last_name'],
                                'email': existing_user['contact_details']['email'],
                                'phone': existing_user['contact_details']['phone']}
                if token:
                    token['user_info'] = user_details
                return 200, "token", token
            else:
                return 400, "Device not found", {}
        else:
            return 400, "Incorrect Email or Password", {}
    else:
        return 400, "Please give login creds", {}
    # else:
    #     return 400, "Device not registered", {}


def logout_user(current_user):
    session_code = current_user['session_code']
    # user_device = mongo.db.user_device.find_one(
    #     {'user_id': ObjectId(current_user['_id']), 'session_code': session_code, "status": "active"})
    if session_code:
        updated_user_device = mongo.db.user.update_one(
            {'_id': ObjectId(current_user['_id']), 'session_code': session_code}, {
                '$set': {'logout_at': datetime.now(), "is_active": False}}).modified_count
        if updated_user_device != 0:
            return 200, "Logout Successfully", {}
        else:
            return 400, "logout Failed", {}
    else:
        return 400, "Device not found", {}


def create_user():
    data = request.json
    if data:
        user_data= mongo.db.user.insert_one(data).inserted_id
        return 200,'Inserted',{}
    else:
        return 400, 'invalid data',{}

# ===============================================================================================================================


def add_device():
    data = request.json
    fingerprint = data.get('fingerprint')
    device_verification_id = None
    existing_device = mongo.db.device.find_one({"fingerprint": fingerprint})
    # If device exists, update its details
    if existing_device:
        # validate payload
        is_validated = schema_validator(data, 'device_model.json')
        if not is_validated:
            # update device table
            mongo.db.device.update_one({"fingerprint": fingerprint}, {"$set": data})
            device_verification_id = ObjectId(existing_device['_id'])
        else:
            return 200, "Validation Error", {}
    # If device doesn't exist, insert new document
    else:
        # validate payload
        is_validated = schema_validator(data, 'device_model.json')
        if not is_validated:
            data['created_at'] = datetime.now()
            data['updated_at'] = datetime.now()
            add_device = mongo.db.device.insert_one(data)
            device_verification_id = str(add_device.inserted_id)

    device_id = {"device_verification_id": str(device_verification_id)}
    return 200, "Device id", device_id


def forgot_password(data):
    email = data.get('email', None)
    # check the user data is there or not
    existing_user = mongo.db.user.find_one({'contact_details.email': email, 'status': 'active'})
    existing_verification_data = mongo.db.verification.find_one(
        {'payload.email': email, 'type': 'signup'})
    if existing_user and existing_verification_data:
        payload = existing_verification_data['payload']
        verification_type = 'forgot_password'
        inserted_id = send_email(verification_type, payload)
        return 200, "verification mail sent", inserted_id
    else:
        return 400, 'User is not existing, Check your email id !', {}


def verify_user():
    data = request.json
    email = data['verification_detail']['email']
    # check for the email validation
    if not is_valid_email(email):
        return 400, "Email is not valid !!!", {}
    phone = data['verification_detail']['phone']
    if not is_valid_phone_number(phone):
        return 400, "Phone number is not valid !!!", {}
    existing_email_lower = mongo.db.user.find_one({'contact_details.email': email.lower()})
    existing_email_upper = mongo.db.user.find_one({'contact_details.email': email.upper()})
    existing_phone = mongo.db.user.find_one({'contact_details.phone': phone})
    if existing_email_lower or existing_email_upper:
        return 400, "Email Id is already exists !!!", {}
    elif existing_phone:
        return 400, "Phone Number is already exists !!!", {}
    else:
        payload = {}
        date_info = {}
        date_info['created_at'] = str(datetime.now())
        date_info['updated_at'] = str(datetime.now())
        payload['device_verification_id'] = data['device_verification_id']
        payload['first_name'] = data['verification_detail']['first_name']
        payload['last_name'] = data['verification_detail']['last_name']
        payload['email'] = email
        payload['phone'] = phone
        payload['phone_code'] = data['verification_detail']['phone_code']
        payload['org_type'] = data['user_role']['org_type']
        if data['user_role']['org_type'] == 'wastelink':
            payload['org_id'] = {}
        else:
            payload['org_id'] = data['user_role']['org_id']
        payload['role'] = data['user_role']['role']

        # Generate verification code and link
        verification_type = 'signup'
        inserted_id = send_email(verification_type, payload)
        return 200, "verification mail sent", inserted_id


def add_user_details(verification_details, password):
    print("==========", verification_details)
    if verification_details:
        contact_details = {}
        data = {}
        date_info = {}
        date_info['created_at'] = str(datetime.now())
        date_info['updated_at'] = str(datetime.now())
        contact_details['first_name'] = verification_details['payload']['first_name']
        contact_details['last_name'] = verification_details['payload']['last_name']
        contact_details['email'] = verification_details['payload']['email']
        contact_details['phone'] = verification_details['payload']['phone']
        data['role'] = verification_details['payload']['role']
        data['org_type'] = verification_details['payload']['org_type']
        if verification_details['payload']['org_type'] == 'wastelink':
            data['org_id'] = {}
        else:
            data['org_id'] = verification_details['payload']['org_id']
        data['contact_details'] = contact_details
        data['email_verified'] = True
        data['phone_verified'] = False
        data['status'] = 'active'
        data['base_model'] = date_info
        email = verification_details['payload']['email']
        existing_email_lower = mongo.db.user.find_one({'contact_details.email': email.lower(), 'status': 'active'})
        existing_email_upper = mongo.db.user.find_one({'contact_details.email': email.upper(), 'status': 'active'})
        existing_phone = mongo.db.user.find_one({'contact_details.email': verification_details['payload']['phone'],
                                                 'status': 'active'})
        if existing_email_upper or existing_email_lower:
            return None, "Email already registered", {}
        elif existing_phone:
            return None, "Phone Number already registered", {}
        else:
            # validate payload
            is_validated = schema_validator(data, 'user_model.json')
            if not is_validated:
                data['password'] = bcrypt.generate_password_hash(password).decode('utf-8')
                data['password_update_on'] = str(datetime.now())
                add_user = mongo.db.user.insert_one(data)
                if add_user:
                    return 200, "Verification Successful", add_user.inserted_id
                else:
                    return 400, "Verification Failed", {}
            else:
                return 400, "Validation Error", {}
    else:
        return 400, "No verification data", {}


def user_device_mapping(user_id, device_verification_id):
    if user_id and device_verification_id:
        date_info = {}
        date_info['created_at'] = str(datetime.now())
        date_info['updated_at'] = str(datetime.now())
        mapping_data = {}
        mapping_data['user_id'] = str(user_id)
        mapping_data['device_id'] = device_verification_id
        mapping_data['status'] = 'active'
        mapping_data['is_active'] = False
        # mapping_data['login_at']=None
        # mapping_data['logout_at']=None
        mapping_data['session_code'] = generate_session_code()
        mapping_data["base_model"] = date_info
        # validate payload
        is_validated = schema_validator(mapping_data, 'userdevice_model.json')
        if not is_validated:
            add_device_mapping = mongo.db.user_device.insert_one(mapping_data)
            if add_device_mapping:
                return add_device_mapping.inserted_id
            else:
                return None
        else:
            return 200, "Validation Error", {}
    else:
        return None


def get_users():
    # page = int(request.args.get('page', 1))
    # per_page = int(request.args.get('per_page', 10))
    #
    # # pagination
    # total_count = mongo.db.user.count_documents({})
    # pages = int(total_count / per_page) + (total_count % per_page > 0)
    # start = (page - 1) * per_page

    # users = mongo.db.user.find({"status": "active"}).skip(start).limit(per_page)
    users = mongo.db.user.find({"status": "active"})
    user_list = []
    if users:
        for user in users:
            user['_id'] = str(user['_id'])
            user_list.append(user)

        # pagination = {
        #     'total': total_count,
        #     'pages': pages,
        #     'page': page,
        #     'per_page': per_page
        # }
        return 200, "User List", user_list, {}
    else:
        return 400, "No Data", {}, {}


def get_user_detail(user_id):
    user = mongo.db.user.find_one({'_id': ObjectId(user_id), 'status': 'active'})
    if user:
        user['_id'] = str(user['_id'])
        return 200, "User Detail", user
    else:
        return 400, "Data Not Found", {}


def user_delete(user_id):
    # Check if user exists in the database
    existing_user = mongo.db.user.find_one({'_id': ObjectId(user_id), 'status': 'active'})
    if not existing_user:
        return 400, "User not found", {}

    # Delete user from the database
    delete_user = mongo.db.user.update_one({'_id': ObjectId(user_id)},
                                           {'$set': {'status': 'inactive', 'deleted_at': datetime.now()}})
    if delete_user.deleted_count == 1:
        delete_device_mapping = mongo.db.user_device.delete_many({'user_id': user_id}, {'status': 'active'})
        return 200, "User Deleted", {}
    else:
        return 200, "User Not Found", {}



def update_user_details(data, user_id):
    first_name = data.get('first_name', None)
    last_name = data.get('last_name', None)
    org_type = data.get('org_type', None)
    org_id = data.get('org_id', None)
    role = data.get('role', None)
    status = data.get('status', None)
    # checking the inputs
    if first_name and last_name and org_type and org_id and role and status:
        # checking the user is existing or not
        existing_user = mongo.db.user.find_one({'_id': ObjectId(user_id)})

        if existing_user:
            update_details = mongo.db.user.find_one({'_id': ObjectId(user_id)}, {"$set": {
                "contact_details.first_name": first_name,
                "contact_details.last_name": last_name,
                "org_type": org_type,
                "org_id": org_id,
                "role": role,
                "status": status
            }})
            if update_details:
                return 200, 'Details updated', {}
            # else:
            #     return 400, 'Details not updated', {}
        else:
            return 400, 'User not found', {}
    else:
        return 400, 'Invalid Input', {}


def user_bulk_upload(data, token):
    token = {'x-access-token': token}
    file = request.files['file']
    if file and '.' in file.filename and file.filename.rsplit('.', 1)[1].lower() not in {'xls', 'xlsx'}:
        return 400, 'Invalid file format. Please upload a valid XLS or XLSX file.', {}

    filename = file.filename
    # Save the uploaded file
    file.save(filename)

    # Open the workbook
    workbook = load_workbook(filename)
    sheet = workbook.active
    is_error = False
    if sheet:
        required_columns = ['NAME', 'EMAIL', 'PHONE', 'ASSOCIATION', 'ASSOCIATION WITH', 'ROLE', 'STATUS']
        header_row = next(sheet.iter_rows(min_row=1, max_row=1))
        header_values = [cell.value for cell in header_row]
        if header_values:
            missing_columns = [column for column in required_columns if column not in header_values]
            if missing_columns:
                error_workbook = Workbook()
                error_sheet = error_workbook.active
                # Write the error message to the error sheet
                error_sheet.append([f'Invalid Headers : {", ".join(missing_columns)}'])
                # Save the error workbook
                error_workbook.save('user_error_rows.xlsx')
                is_error = True
                return 200, f'Invalid Headers : {", ".join(missing_columns)}', {'is_error': is_error}
        else:
            return 400, 'Something went wrong', {}

        # Find rows with errors
        error_rows = []
        correct_rows = []
        for idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):  # Start from row 2 to skip header
            if all(cell.value is None for cell in row):
                # Skip empty rows
                # If it's the last row, don't skip it
                if idx == sheet.max_row:
                    break
                continue

            # Validate the row for errors
            result, error_data = user_data_has_error(row, token)
            if not result:
                error_rows.append(error_data)
            else:
                correct_rows.append(row)
        if len(correct_rows) != 0:
            # Retrieve data from the file when there are no errors
            organisation_data = []
            data = prepare_user_data(correct_rows, data, token)
            if data:
                for org_data in data:
                    # Generate verification code and link
                    verification_type = 'signup'
                    inserted_id = send_email(verification_type, org_data)
                    # return 200, "verification mail sent", inserted_id

            # return 200, 'Data Retrieved Successfully !!!', {}
        if len(error_rows) != 0:
            # # Create a new workbook and sheet
            new_workbook = Workbook()
            new_sheet = new_workbook.active

            # Write the error rows to the new sheet
            header = ['NAME', 'EMAIL', 'PHONE', 'ASSOCIATION', 'ASSOCIATION WITH', 'ROLE', 'STATUS', 'ERROR MESSAGE']
            new_sheet.append(header)  # Add header row

            for cells_data in error_rows:
                cell_value = cells_data[0][0]
                error_message = cells_data[0][1]

                values = [cell.value for cell in cell_value]
                new_sheet.append(values + [error_message])

            # Save the new workbook with errors
            new_filename = 'user_error_rows.xlsx'
            new_workbook.save(new_filename)
            is_error = True
        return 200, 'Uploaded Successfully !!!', {'is_error': is_error}
    else:
        return 400, 'Invalid file'


def user_data_has_error(row, token):
    user_name = row[0].value
    email = row[1].value
    phone = row[2].value
    association = (row[3].value).lower()
    association_with = row[4].value
    role = (row[5].value).lower().replace(" ", "_")
    status = row[6].value

    error_flag = False
    errors = []
    # check for the email validation(regex)
    existing_email_lower = mongo.db.user.find_one({'contact_details.email': email.lower()})
    existing_email_upper = mongo.db.user.find_one({'contact_details.email': email.upper()})
    existing_phone = mongo.db.user.find_one({'contact_details.phone': phone})

    if association != 'wastelink' and not association_with or not isinstance(association_with, str):
        print("Association is missing or not a string.")
        error_flag = True
        errors.append((row, "Association is missing or not a string."))

    if not association or not isinstance(association, str):
        print("Association is missing or not a string.")
        error_flag = True
        errors.append((row, "Pan no is missing or not a string."))

    if not role or not isinstance(role, str):
        print("Role is missing or not a string.")
        error_flag = True
        errors.append((row, "Role is missing or not a string."))

    if association and association_with and role:
        url = os.environ.get('ORG_CODE_URL') + str(association_with)
        response = make_api_call(url=url, method='GET', headers=token)
        if response['status_code'] != 200:
            error_flag = True
            errors.append((row, "Invalid organisation for association"))
        if association == 'auditor agency':
            if role not in ['auditor_manager', 'auditor_user']:
                print("Invalid role for 'auditor agency'")
                error_flag = True
                errors.append((row, 'Invalid role for auditor agency'))
            if response['status_code'] == 200:
                if response['data']['form_type'] == 'auditor':
                    print("Invalid association for auditor agency")
                    error_flag = True
                    errors.append((row, 'Invalid association for auditor agency'))
        elif association in ['pickup location', 'region']:
            if role != 'supplier_user':
                print("Invalid role for 'pickup location' and 'region'")
                error_flag = True
                errors.append((row, "Invalid role for 'pickup location' and 'region'"))
            if response['status_code'] == 200:
                if response['data']['form_type'] not in ['pickup_location', 'zone']:
                    print("Invalid association for pickup location/zone")
                    error_flag = True
                    errors.append((row, 'Invalid association for pickup location/zone'))
        elif association == 'supplier':
            if role not in ['supplier_manager', 'supplier_user']:
                print("Invalid role for supplier")
                error_flag = True
                errors.append((row, "Invalid role for supplier"))
            if response['status_code'] == 200:
                if response['data']['form_type'] != 'supplier':
                    print("Invalid association for supplier")
                    error_flag = True
                    errors.append((row, 'Invalid association for supplier'))
        else:
            print("Invalid association data.It should be any of 'auditor agency','pickup location', 'region','supplier")
            error_flag = True
            errors.append((row,
                           "Invalid association data.It should be any of 'auditor agency','pickup location', 'region','supplier"))
    if association and association == 'wastelink' and role:
        if role not in ['super_admin', 'account_team', 'account_manager', 'processor_executive', 'fpe',
                        'operation_manager', 'support_desk']:
            print("Invalid Role.")
            error_flag = True
            errors.append((row, "Invalid Role."))
        if association_with:
            print("Association with should be empty for 'wastelink'.")
            error_flag = True
            errors.append((row, "Association with should be empty for 'wastelink'."))

    if not user_name or not isinstance(user_name, str):
        print("user name is missing or not a string.")
        error_flag = True
        errors.append((row, "User name is missing or not a string."))

    if not email or not isinstance(email, str):
        print("email is missing or not a string.")
        error_flag = True
        errors.append((row, "Email ID is missing or not a string."))
    if email and not is_valid_email(email):
        print("Email is not valid")
        error_flag = True
        errors.append((row, "Email is not valid."))
    if existing_email_upper is not None and existing_email_lower is not None:
        print("email id is already existing.")
        error_flag = True
        errors.append((row, "Email ID is already existing."))

    if not phone or not isinstance(phone, int):
        print("phone number is missing or not a string.")
        error_flag = True
        errors.append((row, "Phone Number is missing or not a string."))
    if phone and not is_valid_phone_number(phone):
        print("Phone number is not valid")
        error_flag = True
        errors.append((row, "Phone number is not valid."))
    if existing_phone is not None:
        print("phone number is already existing.")
        error_flag = True
        errors.append((row, "Phone Number is already existing."))

    if not status or not isinstance(status, str):
        print("status is missing or not a string.")
        error_flag = True
        errors.append((row, "Status is missing or not a string."))

    if status and status not in ['Active', 'Inactive']:
        error_flag = True
        errors.append((row, "Status must be either 'Active' or 'Inactive'."))
    if error_flag:
        return False, errors
    else:
        return True, None


def prepare_user_data(correct_rows, data, token):
    user_data = []
    for row in correct_rows:  # Start from row 2 to skip header
        payload = {}
        date_info = {}
        date_info['created_at'] = str(datetime.now())
        date_info['updated_at'] = str(datetime.now())
        payload['device_verification_id'] = data
        payload['first_name'] = row[0].value
        payload['last_name'] = row[0].value
        payload['email'] = row[1].value
        payload['phone'] = str(row[2].value)
        payload['phone_code'] = "+91"
        payload['org_type'] = row[3].value
        if row[3].value == 'wastelink':
            payload['org_id'] = []
        else:
            url = os.environ.get('ORG_CODE_URL') + str(row[4].value)
            response = make_api_call(url=url, method='GET', headers=token)
            association_id = response['data']['_id']
            payload['org_id'] = {association_id: {}}
        payload['role'] = row[5].value
        user_data.append(payload)
    return user_data


def get_user_detail_by_email(email_id):
    print("HIIIIIIIIIII")
    user = mongo.db.user.find_one({'contact_details.email': email_id, 'status': 'active'})
    if user:
        print("HEYYYYYYYYY",user)
        user['_id'] = str(user['_id'])

        return 200, "User Detail", user
    else:
        return 400, "Data Not Found", {}